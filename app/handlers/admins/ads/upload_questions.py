import io

import openpyxl
from aiogram import Router, F, Bot
from aiogram.filters import Filter
from aiogram.types import Message, Document, CallbackQuery, InlineKeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder
from sqlalchemy import select, delete, func

from app.filters.is_admin import IsAdmin
from data.config import ADMINS
from database.database import session_maker
from database.models import Question, Test

router = Router()

REQUIRED_COLUMNS = {
    "question",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct",
    "difficulty",
}
VALID_ANSWERS = {"A", "B", "C", "D"}

_pending_upload: dict[int, int] = {}


# =========================================================
# MAXSUS FILTER — faqat test nomi kutilayotganda ishlaydi
# =========================================================


class WaitingTestTitle(Filter):
    async def __call__(self, message: Message) -> bool:
        return _pending_upload.get(message.from_user.id) == -1


# =========================================================
# YORDAMCHI FUNKSIYALAR
# =========================================================


async def get_tests() -> list[Test]:
    async with session_maker() as session:
        result = await session.execute(select(Test).order_by(Test.id))
        return result.scalars().all()


def tests_inline(tests: list[Test], action: str):
    builder = InlineKeyboardBuilder()
    for t in tests:
        icon = "✅" if t.is_active else "❌"
        builder.add(
            InlineKeyboardButton(
                text=f"{icon} {t.title}",
                callback_data=f"{action}:{t.id}",
            )
        )
    builder.adjust(1)
    return builder.as_markup()


def test_main_keyboard():
    builder = InlineKeyboardBuilder()
    builder.add(
        InlineKeyboardButton(text="➕ Test yaratish", callback_data="admin:create_test")
    )
    builder.add(
        InlineKeyboardButton(
            text="📋 Testlar ro'yxati", callback_data="admin:list_tests"
        )
    )
    builder.add(
        InlineKeyboardButton(text="📥 Savol yuklash", callback_data="admin:upload")
    )
    builder.add(InlineKeyboardButton(text="📊 Statistika", callback_data="admin:stat"))
    builder.add(
        InlineKeyboardButton(text="🗑 Savollarni o'chirish", callback_data="admin:clear")
    )
    builder.adjust(1)
    return builder.as_markup()


def back_button(callback_data: str = "admin:menu"):
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="🔙 Orqaga", callback_data=callback_data))
    return builder.as_markup()


# =========================================================
# ADMIN PANEL KIRISH
# =========================================================


@router.message(F.text == "📄 Testlar Bo'limi", IsAdmin(admin_ids=ADMINS))
async def admin_panel(message: Message):
    await message.answer(
        "⚙️ <b>Test panel</b>\n\nNimani qilmoqchisiz?",
        parse_mode="HTML",
        reply_markup=test_main_keyboard(),
    )


@router.callback_query(F.data == "admin:menu", IsAdmin(admin_ids=ADMINS))
async def admin_menu(callback: CallbackQuery):
    await callback.message.edit_text(
        "⚙️ <b>Test panel</b>\n\nNimani qilmoqchisiz?",
        parse_mode="HTML",
        reply_markup=test_main_keyboard(),
    )
    await callback.answer()


# =========================================================
# TEST YARATISH
# =========================================================


@router.callback_query(F.data == "admin:create_test", IsAdmin(admin_ids=ADMINS))
async def create_test_prompt(callback: CallbackQuery):
    _pending_upload[callback.from_user.id] = -1
    await callback.message.edit_text(
        "📝 <b>Yangi test yaratish</b>\n\n"
        "Test nomini yuboring.\n"
        "Masalan: <code>O'tkan Kunlar — Adabiyot Testi</code>",
        parse_mode="HTML",
        reply_markup=back_button(),
    )
    await callback.answer()


# ← WaitingTestTitle filteri: FAQAT _pending_upload[user_id] == -1 bo'lganda ishlaydi
@router.message(F.text, WaitingTestTitle(), IsAdmin(admin_ids=ADMINS))
async def handle_test_title(message: Message):
    title = message.text.strip()
    _pending_upload.pop(message.from_user.id)

    async with session_maker() as session:
        test = Test(title=title, is_active=True)
        session.add(test)
        await session.commit()
        await session.refresh(test)

    await message.answer(
        f"✅ <b>Test yaratildi!</b>\n\n"
        f"📋 Nom: <b>{test.title}</b>\n"
        f"🆔 ID: <code>{test.id}</code>",
        parse_mode="HTML",
        reply_markup=test_main_keyboard(),
    )


# =========================================================
# TESTLAR RO'YXATI
# =========================================================


@router.callback_query(F.data == "admin:list_tests", IsAdmin(admin_ids=ADMINS))
async def list_tests(callback: CallbackQuery):
    tests = await get_tests()

    if not tests:
        await callback.message.edit_text(
            "📭 Hech qanday test yo'q.", reply_markup=back_button()
        )
        await callback.answer()
        return

    async with session_maker() as session:
        lines = ["📋 <b>Testlar ro'yxati:</b>\n"]
        for t in tests:
            q_count = await session.scalar(
                select(func.count(Question.id)).where(
                    Question.test_id == t.id,
                    Question.is_active.is_(True),
                )
            )
            icon = "✅" if t.is_active else "❌"
            lines.append(
                f"{icon} <b>{t.title}</b>\n   🆔 {t.id} | 📝 {q_count} ta savol"
            )

    await callback.message.edit_text(
        "\n".join(lines), parse_mode="HTML", reply_markup=back_button()
    )
    await callback.answer()


# =========================================================
# SAVOL YUKLASH — test tanlash
# =========================================================


@router.callback_query(F.data == "admin:upload", IsAdmin(admin_ids=ADMINS))
async def upload_select_test(callback: CallbackQuery):
    tests = await get_tests()

    if not tests:
        await callback.message.edit_text(
            "📭 Avval test yarating.", reply_markup=back_button()
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        "📥 <b>Savol yuklash</b>\n\nQaysi testga savollar yuklaysiz?",
        parse_mode="HTML",
        reply_markup=tests_inline(tests, "upload"),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("upload:"), IsAdmin(admin_ids=ADMINS))
async def upload_test_selected(callback: CallbackQuery):
    test_id = int(callback.data.split(":")[1])

    async with session_maker() as session:
        result = await session.execute(select(Test).where(Test.id == test_id))
        test = result.scalar_one_or_none()

    if not test:
        await callback.answer("Test topilmadi.", show_alert=True)
        return

    _pending_upload[callback.from_user.id] = test_id

    await callback.message.edit_text(
        f"📥 <b>{test.title}</b>\n\n"
        "Endi <b>.xlsx</b> faylni yuboring.\n\n"
        "<b>Ustunlar:</b>\n"
        "• <code>question</code> — savol matni\n"
        "• <code>option_a / b / c / d</code> — variantlar\n"
        "• <code>correct</code> — to'g'ri javob (A/B/C/D)\n"
        "• <code>difficulty</code> — qiyinlik (-1.5, 0.0, 2.3...)\n\n"
        "⚠️ 1-qator sarlavha bo'lishi shart.",
        parse_mode="HTML",
        reply_markup=back_button("admin:upload"),
    )
    await callback.answer()


# =========================================================
# EXCEL FAYLNI QABUL QILISH
# =========================================================


@router.message(F.document, IsAdmin(admin_ids=ADMINS))
async def handle_excel_upload(message: Message, bot: Bot):
    doc: Document = message.document

    if not doc.file_name or not doc.file_name.endswith(".xlsx"):
        await message.answer(
            "⚠️ Faqat <b>.xlsx</b> fayl qabul qilinadi.", parse_mode="HTML"
        )
        return

    test_id = _pending_upload.pop(message.from_user.id, None)
    if test_id is None:
        await message.answer(
            "⚠️ Avval testni tanlang.", reply_markup=test_main_keyboard()
        )
        return

    async with session_maker() as session:
        result = await session.execute(select(Test).where(Test.id == test_id))
        test = result.scalar_one_or_none()

    if not test:
        await message.answer("❌ Test topilmadi.")
        return

    status_msg = await message.answer(
        f"⏳ <b>{test.title}</b> — savollar yuklanmoqda...", parse_mode="HTML"
    )

    file = await bot.get_file(doc.file_id)
    file_bytes = await bot.download_file(file.file_path)
    content = file_bytes.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        await status_msg.edit_text(
            "❌ Faylni o'qib bo'lmadi. Buzilgan bo'lishi mumkin."
        )
        return

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        await status_msg.edit_text(
            f"❌ Ustunlar topilmadi: <code>{', '.join(missing)}</code>",
            parse_mode="HTML",
        )
        return

    col = {name: idx for idx, name in enumerate(headers)}
    questions_to_add = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        def get(field):
            idx = col.get(field)
            return (
                str(row[idx]).strip()
                if idx is not None and row[idx] is not None
                else ""
            )

        q_text = get("question")
        opt_a = get("option_a")
        opt_b = get("option_b")
        opt_c = get("option_c")
        opt_d = get("option_d")
        correct = get("correct").upper()
        diff_raw = get("difficulty")

        row_errors = []
        if not q_text:
            row_errors.append("savol bo'sh")
        if not all([opt_a, opt_b, opt_c, opt_d]):
            row_errors.append("variant bo'sh")
        if correct not in VALID_ANSWERS:
            row_errors.append(f"correct='{correct}'")
        try:
            difficulty = float(diff_raw)
        except ValueError:
            row_errors.append(f"difficulty='{diff_raw}'")
            difficulty = 0.0

        if row_errors:
            errors.append(f"Qator {row_num}: {'; '.join(row_errors)}")
            continue

        questions_to_add.append(
            Question(
                test_id=test_id,
                text=q_text,
                option_a=opt_a,
                option_b=opt_b,
                option_c=opt_c,
                option_d=opt_d,
                correct=correct,
                difficulty=difficulty,
                is_active=True,
            )
        )

    if not questions_to_add:
        await status_msg.edit_text(
            "❌ Yaroqli savollar topilmadi.\n\n" + "\n".join(errors[:10]),
            parse_mode="HTML",
        )
        return

    async with session_maker() as session:
        session.add_all(questions_to_add)
        await session.commit()
        total = await session.scalar(
            select(func.count(Question.id)).where(
                Question.test_id == test_id,
                Question.is_active.is_(True),
            )
        )

    text = (
        f"✅ <b>{len(questions_to_add)} ta savol yuklandi!</b>\n\n"
        f"📋 Test: <b>{test.title}</b>\n"
        f"📦 Testdagi jami faol savollar: {total}\n"
    )
    if errors:
        text += f"\n⚠️ {len(errors)} ta qator o'tkazib yuborildi:\n" + "\n".join(
            errors[:5]
        )
        if len(errors) > 5:
            text += f"\n... va yana {len(errors) - 5} ta"

    await status_msg.edit_text(
        text, parse_mode="HTML", reply_markup=test_main_keyboard()
    )


# =========================================================
# STATISTIKA
# =========================================================


@router.callback_query(F.data == "admin:stat", IsAdmin(admin_ids=ADMINS))
async def show_stat(callback: CallbackQuery):
    async with session_maker() as session:
        tests_result = await session.execute(select(Test).order_by(Test.id))
        tests = tests_result.scalars().all()

        lines = ["📊 <b>Savollar statistikasi</b>\n"]
        grand_total = 0

        for t in tests:
            active = await session.scalar(
                select(func.count(Question.id)).where(
                    Question.test_id == t.id, Question.is_active.is_(True)
                )
            )
            total = await session.scalar(
                select(func.count(Question.id)).where(Question.test_id == t.id)
            )
            grand_total += total
            lines.append(
                f"📋 <b>{t.title}</b>\n"
                f"   Jami: {total} | Faol: {active} | Nofaol: {total - active}\n"
            )
        lines.append(f"\n🔢 Umumiy: {grand_total} ta")

    await callback.message.edit_text(
        "\n".join(lines), parse_mode="HTML", reply_markup=back_button()
    )
    await callback.answer()


# =========================================================
# SAVOLLARNI O'CHIRISH
# =========================================================


@router.callback_query(F.data == "admin:clear", IsAdmin(admin_ids=ADMINS))
async def clear_select_test(callback: CallbackQuery):
    tests = await get_tests()

    if not tests:
        await callback.message.edit_text(
            "📭 Hech qanday test yo'q.", reply_markup=back_button()
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        "🗑 <b>Savollarni o'chirish</b>\n\nQaysi testni tozalaysiz?",
        parse_mode="HTML",
        reply_markup=tests_inline(tests, "clear"),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("clear:"), IsAdmin(admin_ids=ADMINS))
async def clear_test_selected(callback: CallbackQuery):
    test_id = int(callback.data.split(":")[1])

    async with session_maker() as session:
        result = await session.execute(select(Test).where(Test.id == test_id))
        test = result.scalar_one_or_none()
        if not test:
            await callback.answer("Test topilmadi.", show_alert=True)
            return
        count = await session.scalar(
            select(func.count(Question.id)).where(Question.test_id == test_id)
        )

    builder = InlineKeyboardBuilder()
    builder.add(
        InlineKeyboardButton(
            text="✅ Ha, o'chirish", callback_data=f"clear_ok:{test_id}"
        )
    )
    builder.add(
        InlineKeyboardButton(text="❌ Bekor qilish", callback_data="admin:clear")
    )
    builder.adjust(2)

    await callback.message.edit_text(
        f"⚠️ <b>{test.title}</b>\n\n"
        f"<b>{count} ta savol</b> o'chiriladi. Ishonchingiz komilmi?",
        parse_mode="HTML",
        reply_markup=builder.as_markup(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("clear_ok:"), IsAdmin(admin_ids=ADMINS))
async def clear_confirmed(callback: CallbackQuery):
    test_id = int(callback.data.split(":")[1])

    async with session_maker() as session:
        result = await session.execute(select(Test).where(Test.id == test_id))
        test = result.scalar_one_or_none()
        await session.execute(delete(Question).where(Question.test_id == test_id))
        await session.commit()

    await callback.message.edit_text(
        f"🗑 <b>{test.title}</b> — barcha savollar o'chirildi.",
        parse_mode="HTML",
        reply_markup=test_main_keyboard(),
    )
    await callback.answer()
