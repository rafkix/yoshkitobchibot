from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from aiogram import Bot, Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile, ReplyKeyboardRemove
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder

from app.filters.is_admin import IsAdmin
from app.keyboards.admin_flow import CANCEL_TEXT, cancel_reply_keyboard
from app.keyboards.reply import admin_menu
from data.config import ADMINS
from database.database import session_maker
from database.models import DirectionType, ContestType
from database.services.user_service import UserService

router = Router()

# =========================================================
# CONSTANTS & CONFIG
# =========================================================

DIRECTION_LABELS = {
    DirectionType.AGE_10_14: "10-14 yosh (2012-2016)",
    DirectionType.AGE_15_19: "15-19 yosh (2007-2011)",
    DirectionType.AGE_20_30: "20-30 yosh (1996-2006)",
}

CONTEST_LABELS = {
    ContestType.YOSH_KITOBXON_2026: "“Yosh kitobxon” 2026",
}

PAGE_SIZE = 10


class UserSearchState(StatesGroup):
    waiting_query = State()
    waiting_score = State()


# =========================================================
# KEYBOARDS (CHIROYLI VA TARTIBLI VERSIYASI)
# =========================================================


def users_list_keyboard(
    users: list,
    page: int,
    total: int,
    filter_mode: str = "all",
):
    kb = InlineKeyboardBuilder()

    for user in users:
        icon = "🟢" if user.is_registered else "🟡"

        name = (user.full_name or "Noma‘lum")[:18]

        kb.button(
            text=f"{icon} {name} • {user.total_score}b",
            callback_data=f"admuv:{user.user_id}:{page}:{filter_mode}",
        )

    pages = max(1, (total - 1) // PAGE_SIZE + 1)

    kb.button(
        text="⬅️",
        callback_data=f"admup:{max(page - 1, 0)}:{filter_mode}",
    )

    kb.button(
        text=f"{page + 1}/{pages}",
        callback_data="admup_noop",
    )

    kb.button(
        text="➡️",
        callback_data=f"admup:{min(page + 1, pages - 1)}:{filter_mode}",
    )

    kb.button(
        text="👥 Barchasi",
        callback_data="admuf:all:0",
    )

    kb.button(
        text="🟢 Aktiv",
        callback_data="admuf:reg:0",
    )

    kb.button(
        text="🟡 Chala",
        callback_data="admuf:unreg:0",
    )

    kb.button(
        text="🔍 Qidiruv",
        callback_data="admus",
    )

    kb.button(
        text="📥 Excel",
        callback_data=f"admux:{filter_mode}",
    )

    kb.button(
        text="🔄 Yangilash",
        callback_data=f"admup:{page}:{filter_mode}",
    )

    kb.button(
        text="🏠 Bosh menyu",
        callback_data="admub",
    )

    kb.adjust(
        1,
        1,
        1,
        3,
        3,
        1,
    )

    return kb.as_markup()


def user_detail_keyboard(
    user_id: int,
    back_page: int,
    filter_mode: str,
    is_registered: bool,
):
    kb = InlineKeyboardBuilder()

    kb.button(
        text="✉️ Xabar yuborish",
        callback_data=f"admum:{user_id}",
    )

    kb.button(
        text="🎯 Ball tahrirlash",
        callback_data=f"admubs:{user_id}:{back_page}:{filter_mode}",
    )

    status_text = "🔴 Deaktivatsiya" if is_registered else "🟢 Aktivlashtirish"

    kb.button(
        text=status_text,
        callback_data=f"admust:{user_id}:{back_page}:{filter_mode}",
    )

    kb.button(
        text="🗑 O'chirish",
        callback_data=f"admudel:{user_id}:{back_page}:{filter_mode}",
    )

    kb.button(
        text="📊 Statistika",
        callback_data=f"admustat:{user_id}",
    )

    kb.button(
        text="👥 Referallar",
        callback_data=f"admuref:{user_id}",
    )

    kb.button(
        text="⬅️ Ortga",
        callback_data=f"admup:{back_page}:{filter_mode}",
    )

    kb.adjust(
        2,
        2,
        2,
        1,
    )

    return kb.as_markup()


def confirm_delete_keyboard(user_id: int, back_page: int = 0, filter_mode: str = "all"):
    """O'chirishni tasdiqlash paneli"""
    builder = InlineKeyboardBuilder()
    builder.button(
        text="🛑 Ha, butunlay o‘chirilsin",
        callback_data=f"admudelok:{user_id}:{back_page}:{filter_mode}",
    )
    builder.button(
        text="🟢 Yo‘q, bekor qilish",
        callback_data=f"admuv:{user_id}:{back_page}:{filter_mode}",
    )
    builder.adjust(1)
    return builder.as_markup()


# =========================================================
# HELPERS
# =========================================================


def user_detail_text(user, referrals_total: int, referrals_registered: int) -> str:
    direction = DIRECTION_LABELS.get(user.direction, "⚠️ Belgilanmagan")
    contest = CONTEST_LABELS.get(user.contest, "⚠️ Belgilanmagan")
    status = "🟢 Ro‘yxatdan o‘tgan" if user.is_registered else "🟡 Ro‘yxatdan o‘tmagan"
    ref_by = (
        f"<code>{user.referred_by}</code>" if user.referred_by else "<i>Hech kim</i>"
    )

    return (
        "<b>👤 FOYDALANUVCHINING TO'LIQ PROFILI</b>\n"
        "📂 <i>ID:</i> <code>{user.user_id}</code>\n"
        "───────────────────\n"
        f"📋 <b>Shaxsiy ma‘lumotlar:</b>\n"
        f" 👤 <b>F.I.Sh:</b> {user.full_name or '—'}\n"
        f" 📅 <b>Tug'ilgan sana:</b> {user.birth_date or '—'}\n"
        f" 📞 <b>Telefon:</b> <code>{user.phone_number or '—'}</code>\n\n"
        f"📍 <b>Manzil ma‘lumotlari:</b>\n"
        f" 🗺 <b>Viloyat:</b> {user.region or '—'}\n"
        f" 🏙 <b>Tuman:</b> {user.district or '—'}\n"
        f" 🏡 <b>Mahalla:</b> {user.neighborhood or '—'}\n"
        f" 🏫 <b>Ish/O'qish:</b> {user.workplace or '—'}\n\n"
        f"🏆 <b>Tanlov holati:</b>\n"
        f" 🎮 <b>Tanlov:</b> {contest}\n"
        f" 📊 <b>Toifa:</b> {direction}\n\n"
        f"📈 <b>Ballar balansi:</b>\n"
        f" 🚀 <b>Umumiy ball:</b> {user.total_score} ball\n"
        f" 📝 Test ballari: {user.test_score} b\n"
        f" 👥 Taklif ballari: {user.referral_score} b\n\n"
        f"👥 <b>Referal tizimi:</b>\n"
        f" 🔗 <b>Taklif qildi:</b> {ref_by}\n"
        f" 🎯 <b>Jami takliflar:</b> {referrals_total} ta\n"
        f" ↳ 🎉 Tasdiqlanganlar: {referrals_registered} ta\n\n"
        f"📌 <b>Holat:</b> {status}\n"
        f"📅 <b>Tizimga kirdi:</b> {user.created_at.strftime('%Y-%m-%d %H:%M')}\n"
        "───────────────────"
    )


async def get_filtered_users(service: UserService, filter_mode: str) -> list:
    if filter_mode == "reg":
        return await service.get_all_users(registered_only=True)
    elif filter_mode == "unreg":
        all_users = await service.get_all_users()
        return [u for u in all_users if not u.is_registered]
    return await service.get_all_users()


def build_list_text(
    users: list, page: int, total: int, filter_mode: str, search_query: str = ""
) -> str:
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)
    filter_names = {
        "all": "Barchasi",
        "reg": "🟢 Faqat Ro‘yxatdan o‘tganlar",
        "unreg": "🟡 Faqat Ro‘yxatdan o‘tmaganlar",
    }
    header = filter_names.get(filter_mode, "Barchasi")
    text = f"👥 <b>Foydalanuvchilar — [{header}]</b>\n\n📊 Tizimdagi jami: <b>{total}</b> ta"
    if total > 0:
        text += f" | Ko‘rsatilmoqda: <b>{start + 1}–{end}</b>"
    if search_query:
        text += f"\n🔍 Qidiruv so'rovi: <code>{search_query}</code>"
    return text


# =========================================================
# HANDLERS (LOGIKA VA XAVFSIZLIK INTEGRATSIYASI)
# =========================================================


@router.message(F.text == "👥 Foydalanuvchilar", IsAdmin(admin_ids=ADMINS))
async def admin_users_menu(message: Message, state: FSMContext):
    await state.clear()
    async with session_maker() as session:
        service = UserService(session)
        all_users = await service.get_all_users()

    total = len(all_users)
    reg = sum(1 for u in all_users if u.is_registered)
    unreg = total - reg
    page_users = all_users[:PAGE_SIZE]

    text = (
        "👥 <b>Foydalanuvchilarni boshqarish markazi</b>\n\n"
        f"📊 <b>Jami foydalanuvchilar:</b> {total} ta\n"
        f" 🟢 Ro‘yxatdan o‘tganlar: <b>{reg}</b> ta\n"
        f" 🟡 Chala ro‘yxatdagilar: <b>{unreg}</b> ta\n\n"
        "👇 Quyidagi ro'yxatdan kerakli foydalanuvchini tanlang:"
    )

    await message.answer(
        text,
        parse_mode="HTML",
        reply_markup=users_list_keyboard(page_users, 0, total, "all"),
    )


@router.callback_query(F.data.startswith("admuf:"), IsAdmin(admin_ids=ADMINS))
@router.callback_query(F.data.startswith("admup:"), IsAdmin(admin_ids=ADMINS))
async def admin_users_navigation(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    # Bu handler ham filtrlash, ham sahifalashni xavfsiz boshqaradi
    if parts[0] == "admuf":
        filter_mode = parts[1]
        page = 0
    else:
        page = int(parts[1])
        filter_mode = parts[2] if len(parts) > 2 else "all"

    state_data = await state.get_data()
    search_query = state_data.get("current_search_query", "")

    async with session_maker() as session:
        service = UserService(session)
        if search_query:
            users = await service.search_users(search_query)
            if filter_mode == "reg":
                users = [u for u in users if u.is_registered]
            elif filter_mode == "unreg":
                users = [u for u in users if not u.is_registered]
        else:
            users = await get_filtered_users(service, filter_mode)

    total = len(users)
    page_users = users[page * PAGE_SIZE : (page + 1) * PAGE_SIZE]

    await callback.message.edit_text(
        build_list_text(users, page, total, filter_mode, search_query),
        parse_mode="HTML",
        reply_markup=users_list_keyboard(page_users, page, total, filter_mode),
    )
    await callback.answer()


@router.callback_query(F.data == "admup_noop")
async def noop(callback: CallbackQuery):
    await callback.answer()


@router.callback_query(F.data.startswith("admuv:"), IsAdmin(admin_ids=ADMINS))
async def admin_user_detail(callback: CallbackQuery):
    parts = callback.data.split(":")
    user_id = int(parts[1])
    back_page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
    filter_mode = parts[3] if len(parts) > 3 else "all"

    async with session_maker() as session:
        service = UserService(session)
        user = await service.get_user(user_id)
        if not user:
            return await callback.answer(
                "❌ Xatolik: Foydalanuvchi topilmadi", show_alert=True
            )

        referrals_total = await service.get_referrals_count(user_id)
        referrals_registered = await service.get_registered_referrals_count(user_id)

    await callback.message.edit_text(
        user_detail_text(user, referrals_total, referrals_registered),
        parse_mode="HTML",
        reply_markup=user_detail_keyboard(
            user_id, back_page, filter_mode, user.is_registered
        ),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admust:"), IsAdmin(admin_ids=ADMINS))
async def admin_toggle_status(callback: CallbackQuery):
    parts = callback.data.split(":")
    user_id = int(parts[1])
    back_page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
    filter_mode = parts[3] if len(parts) > 3 else "all"

    async with session_maker() as session:
        service = UserService(session)
        user = await service.get_user(user_id)
        if not user:
            return await callback.answer("❌ Foydalanuvchi topilmadi", show_alert=True)

        new_status = not user.is_registered
        await service.update_user(user_id, is_registered=new_status)

        referrals_total = await service.get_referrals_count(user_id)
        referrals_registered = await service.get_registered_referrals_count(user_id)
        user.is_registered = new_status

    await callback.answer("✅ Status muvaffaqiyatli o'zgartirildi!", show_alert=True)
    await callback.message.edit_text(
        user_detail_text(user, referrals_total, referrals_registered),
        parse_mode="HTML",
        reply_markup=user_detail_keyboard(user_id, back_page, filter_mode, new_status),
    )


@router.callback_query(F.data.startswith("admudel:"), IsAdmin(admin_ids=ADMINS))
async def admin_delete_confirm(callback: CallbackQuery):
    parts = callback.data.split(":")
    user_id = int(parts[1])
    back_page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
    filter_mode = parts[3] if len(parts) > 3 else "all"

    await callback.message.edit_text(
        f"⚠️ <b>DIQQAT! Foydalanuvchini o'chirish</b>\n\n"
        f"Rostdan ham ID: <code>{user_id}</code> bo'lgan foydalanuvchini bazadan butunlay o'chirmoqchimisiz?\n"
        f"⚠️ <i>Bu amalni aslo orqaga qaytarib bo'lmaydi!</i>",
        parse_mode="HTML",
        reply_markup=confirm_delete_keyboard(user_id, back_page, filter_mode),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admudelok:"), IsAdmin(admin_ids=ADMINS))
async def admin_delete_ok(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    user_id = int(parts[1])
    back_page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
    filter_mode = parts[3] if len(parts) > 3 else "all"

    state_data = await state.get_data()
    search_query = state_data.get("current_search_query", "")

    async with session_maker() as session:
        service = UserService(session)
        deleted = await service.delete_user(user_id)

    if not deleted:
        return await callback.answer(
            "❌ O'chirish jarayonida xatolik yuz berdi", show_alert=True
        )

    await callback.answer("🗑 Foydalanuvchi muvaffaqiyatli o'chirildi", show_alert=True)

    async with session_maker() as session:
        service = UserService(session)
        if search_query:
            users = await service.search_users(search_query)
            if filter_mode == "reg":
                users = [u for u in users if u.is_registered]
            elif filter_mode == "unreg":
                users = [u for u in users if not u.is_registered]
        else:
            users = await get_filtered_users(service, filter_mode)

    total = len(users)
    page = min(back_page, max(0, (total - 1) // PAGE_SIZE))
    page_users = users[page * PAGE_SIZE : (page + 1) * PAGE_SIZE]

    await callback.message.edit_text(
        build_list_text(users, page, total, filter_mode, search_query),
        parse_mode="HTML",
        reply_markup=users_list_keyboard(page_users, page, total, filter_mode),
    )


@router.callback_query(F.data.startswith("admubs:"), IsAdmin(admin_ids=ADMINS))
async def admin_edit_score_start(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    user_id = int(parts[1])
    back_page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
    filter_mode = parts[3] if len(parts) > 3 else "all"

    async with session_maker() as session:
        service = UserService(session)
        user = await service.get_user(user_id)
        if not user:
            return await callback.answer("❌ Foydalanuvchi topilmadi", show_alert=True)

    await state.update_data(
        score_target_user_id=user_id,
        score_back_page=back_page,
        score_filter_mode=filter_mode,
    )
    await state.set_state(UserSearchState.waiting_score)

    await callback.message.answer(
        f"🎯 <b>Referal ballni o'zgartirish</b>\n\n"
        f"Foydalanuvchi ID: <code>{user_id}</code>\n"
        f"Joriy ball miqdori: <b>{user.referral_score} b</b>\n\n"
        "📝 Yangi referal ball qiymatini kiriting:",
        parse_mode="HTML",
        reply_markup=cancel_reply_keyboard(),
    )
    await callback.answer()


@router.message(UserSearchState.waiting_score, IsAdmin(admin_ids=ADMINS))
async def admin_edit_score_save(message: Message, state: FSMContext):
    if message.text and message.text.strip() in {"/cancel", CANCEL_TEXT}:
        await state.clear()
        return await message.answer(
            "❌ Jarayon bekor qilindi.", reply_markup=admin_menu()
        )

    if not message.text or not message.text.strip().lstrip("-").isdigit():
        return await message.answer(
            "❌ Faqat butun son kiriting (Masalan: 10 yoki -5):"
        )

    new_score = int(message.text.strip())
    data = await state.get_data()
    user_id = data["score_target_user_id"]

    await state.clear()

    async with session_maker() as session:
        service = UserService(session)
        # Sening Servisingdagi to'g'ri metodga ulandi
        user = await service.set_referral_score(user_id, new_score)

    if not user:
        return await message.answer(
            "❌ Tizim xatosi: Foydalanuvchi topilmadi.", reply_markup=admin_menu()
        )

    await message.answer(
        f"✅ ID: <code>{user_id}</code> foydalanuvchining referal bali <b>{new_score}</b> qilib o'rnatildi.\n"
        f"📊 Yangi jami ball: <b>{user.total_score}</b>",
        parse_mode="HTML",
        reply_markup=admin_menu(),
    )


@router.callback_query(F.data.startswith("admux:"), IsAdmin(admin_ids=ADMINS))
async def admin_export_excel(callback: CallbackQuery, bot: Bot, state: FSMContext):
    filter_mode = callback.data.split(":")[1] if ":" in callback.data else "all"
    state_data = await state.get_data()
    search_query = state_data.get("current_search_query", "")

    await callback.answer("⏳ Excel fayl shakllantirilmoqda, iltimos kuting...")

    async with session_maker() as session:
        service = UserService(session)
        if search_query:
            users = await service.search_users(search_query)
            if filter_mode == "reg":
                users = [u for u in users if u.is_registered]
            elif filter_mode == "unreg":
                users = [u for u in users if not u.is_registered]
        else:
            users = await get_filtered_users(service, filter_mode)

    if not users:
        return await callback.message.answer(
            "❌ Eksport qilish uchun foydalanuvchilar mavjud emas."
        )

    # Excel builder funksiyangiz o'zgarmasdan qoldi
    buf = build_excel(users)
    filename = f"users_{filter_mode}_{datetime.now().strftime('%d_%m')}.xlsx"

    await bot.send_document(
        chat_id=callback.from_user.id,
        document=BufferedInputFile(buf.read(), filename=filename),
        caption=f"📥 <b>Eksport yakunlandi!</b>\n📊 Jami satrlar: {len(users)} ta",
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("admum:"), IsAdmin(admin_ids=ADMINS))
async def admin_message_to_user_start(callback: CallbackQuery, state: FSMContext):
    user_id = int(callback.data.split(":")[1])
    await state.update_data(message_target_user_id=user_id)
    await state.set_state(UserSearchState.waiting_query)
    await callback.message.answer(
        f"✉️ <b>Foydalanuvchiga shaxsiy xabar yuborish</b>\n\n"
        f"Kimga: <code>{user_id}</code>\n\n"
        f"✍️ Xabar matnini kiriting (HTML format qo'llab-quvvatlanadi):",
        parse_mode="HTML",
        reply_markup=cancel_reply_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "admus", IsAdmin(admin_ids=ADMINS))
async def admin_users_search_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(UserSearchState.waiting_query)
    await state.update_data(message_target_user_id=None)
    await callback.message.answer(
        "🔍 <b>Foydalanuvchi qidiruv tizimi</b>\n\n"
        "Quyidagi parametrlardan birini kiriting:\n"
        "• 📝 Ism-familiya (yoki qismi)\n"
        "• 📞 Telefon raqami\n"
        "• 🆔 Telegram unikal ID raqami",
        parse_mode="HTML",
        reply_markup=cancel_reply_keyboard(),
    )
    await callback.answer()


@router.message(UserSearchState.waiting_query, IsAdmin(admin_ids=ADMINS))
async def admin_users_search_or_message(message: Message, state: FSMContext):
    if message.text and message.text.strip() in {"/cancel", CANCEL_TEXT}:
        await state.clear()
        return await message.answer(
            "❌ Amaliyot bekor qilindi.", reply_markup=admin_menu()
        )

    data = await state.get_data()
    target_user_id = data.get("message_target_user_id")

    # Xabar yuborish rejimi faolligini tekshirish
    if target_user_id:
        await state.clear()
        try:
            await message.bot.send_message(
                chat_id=target_user_id,
                text=f"📩 <b>Tizim ma‘muriyatidan xabar:</b>\n\n{message.text}",
                parse_mode="HTML",
            )
            await message.answer(
                f"✅ Xabar <code>{target_user_id}</code> ga muvaffaqiyatli yetkazildi.",
                parse_mode="HTML",
                reply_markup=admin_menu(),
            )
        except Exception as e:
            await message.answer(
                f"❌ Xabarni yetkazib bo'lmadi: {e}", reply_markup=admin_menu()
            )
        return

    # Qidiruv rejimi logikasi
    query = message.text.strip()
    await state.clear()
    await state.update_data(current_search_query=query)

    async with session_maker() as session:
        service = UserService(session)
        if query.isdigit():
            user = await service.get_user(int(query))
            users = [user] if user else []
        else:
            users = await service.search_users(query)

    total = len(users)
    if not users:
        return await message.answer(
            f"🤷‍♂️ Kechirasiz, <b>'{query}'</b> so'rovi bo'yicha hech qanday ma‘lumot topilmadi.",
            parse_mode="HTML",
            reply_markup=admin_menu(),
        )

    await message.answer(
        f"🔍 <b>Qidiruv natijalari:</b> {query}\n📊 Topilganlar soni: <b>{total}</b> ta",
        parse_mode="HTML",
        reply_markup=users_list_keyboard(users[:PAGE_SIZE], 0, total, "all"),
    )


@router.callback_query(F.data == "admub", IsAdmin(admin_ids=ADMINS))
async def admin_users_back(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    await callback.message.answer(
        "🔐 Bosh adminstratorlik paneli. Kerakli menyuni tanlang.",
        reply_markup=admin_menu(),
    )
    await callback.answer()


def build_excel(users: list) -> BytesIO:
    # Berilgan mukammal Excel yaratish funksiyangiz o'z holaticha integratsiya qilindi
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers = [
        "№",
        "Telegram ID",
        "F.I.Sh.",
        "Tug‘ilgan sana",
        "Viloyat",
        "Tuman",
        "Mahalla",
        "Ish/o‘qish joyi",
        "Telefon",
        "Tanlov",
        "Yo‘nalish",
        "Test ball",
        "Referal ball",
        "Umumiy ball",
        "Holat",
        "Kim taklif qilgan",
        "Qo‘shilgan sana",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    for i, u in enumerate(users, 1):
        direction = DIRECTION_LABELS.get(u.direction, u.direction or "—")
        contest = CONTEST_LABELS.get(u.contest, u.contest or "—")
        status = "Ro‘yxatdan o‘tgan" if u.is_registered else "o‘tmagan"
        created = u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "—"

        row = [
            i,
            u.user_id,
            u.full_name or "—",
            str(u.birth_date) if u.birth_date else "—",
            u.region or "—",
            u.district or "—",
            u.neighborhood or "—",
            u.workplace or "—",
            u.phone_number or "—",
            contest,
            direction,
            u.test_score,
            u.referral_score,
            u.total_score,
            status,
            str(u.referred_by) if u.referred_by else "—",
            created,
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i + 1, column=col, value=val)

        if i % 2 == 0:
            fill = PatternFill("solid", fgColor="EBF3FB")
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = fill

    col_widths = [4, 14, 30, 14, 16, 16, 20, 25, 16, 20, 18, 10, 12, 12, 16, 16, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
